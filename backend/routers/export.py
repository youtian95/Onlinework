from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body
from fastapi.responses import Response, StreamingResponse, JSONResponse
from sqlmodel import Session, select, delete
from backend.db import engine
from backend.models import (
    Student,
    Attempt,
    ProblemState,
    ProblemSubmission,
    SystemSetting,
    TeamWorkConfig,
    Team,
    TeamMember,
    TeamSubproblemClaim,
    TeamAttempt,
    TeamSubmission,
)
from backend.core.config import ARCHIVES_DIR, PUBLIC_DIR
from backend.core.security import verify_token
from backend.services.problems import (
    load_problem_script,
    PROBLEMS_DIR,
    get_problem_subproblem_bundle,
    get_problem_input_ids,
    get_stable_rng,
    ensure_meta_inputs,
    DEFAULT_INPUT_SCORE
)
from jinja2 import Template
import csv
import io
import zipfile
import markdown
import os
from datetime import datetime, timezone
from typing import Optional, List
from pydantic import BaseModel

# ReportLab imports for manual PDF construction
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import logging
import shutil

router = APIRouter(prefix="/admin/export", tags=["export"])

def get_session():
    with Session(engine) as session:
        yield session

def verify_admin_token(token: str):
    if not token:
        raise HTTPException(status_code=401, detail="Missing Token")
    try:
        # 如果是 Bearer Token，去掉前缀
        clean_token = token.replace("Bearer ", "") if "Bearer " in token else token
        payload = verify_token(clean_token)
        if not payload or payload.get("role") != "admin":
            raise HTTPException(status_code=401, detail="Unauthorized")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid Token")

class ArchiveCreateRequest(BaseModel):
    name: str

class ExportInputHelper:
    def __init__(self, attempts_map, meta_inputs):
        self.inputs = []
        self.attempts_map = attempts_map
        self.meta_inputs = meta_inputs
        self.total_possible = 0
        self.total_obtained = 0

    def __call__(self, input_id):
        self.inputs.append(input_id)
        attempt = self.attempts_map.get(input_id)
        
        config = self.meta_inputs.get(input_id, {})
        max_score = config.get("score", DEFAULT_INPUT_SCORE)
        self.total_possible += max_score
        
        user_score = 0
        ans_text = "____"
        is_correct = False
        
        if attempt:
            ans_text = attempt.last_answer if attempt.last_answer is not None else "____"
            if attempt.correct:
                user_score = max_score
                is_correct = True
        
        self.total_obtained += user_score
        
        color = "#67c23a" if is_correct else "#f56c6c"
        return f'<span style="color: {color}">**{ans_text}**</span> <sub>({user_score}/{max_score})</sub>'

# --- Helper Functions for Dump/Restore ---

def create_db_dump_bytes(session: Session) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        
        # 1. Students
        s_out = io.StringIO()
        s_writer = csv.writer(s_out)
        s_writer.writerow(["id", "student_id", "name", "enabled", "is_test", "is_deleted", "password_hash"]) # Header
        students = session.exec(select(Student)).all()
        for s in students:
            s_writer.writerow([s.id, s.student_id, s.name, s.enabled, s.is_test, s.is_deleted, s.password_hash])
        zip_file.writestr("students.csv", s_out.getvalue().encode("utf-8-sig"))
        
        # 2. Attempts
        a_out = io.StringIO()
        a_writer = csv.writer(a_out)
        a_writer.writerow(["id", "student_id", "problem_id", "input_id", "attempts", "correct", "last_answer", "updated_at"])
        attempts = session.exec(select(Attempt)).all()
        for a in attempts:
            a_writer.writerow([a.id, a.student_id, a.problem_id, a.input_id, a.attempts, a.correct, a.last_answer, a.updated_at.isoformat() if a.updated_at else ""])
        zip_file.writestr("attempts.csv", a_out.getvalue().encode("utf-8-sig"))

        # 3. Problem Submissions
        ps_out = io.StringIO()
        ps_writer = csv.writer(ps_out)
        ps_writer.writerow(["id", "student_id", "problem_id", "pdf_path", "original_filename", "updated_at"])
        problem_submissions = session.exec(select(ProblemSubmission)).all()
        for p in problem_submissions:
            ps_writer.writerow([
                p.id,
                p.student_id,
                p.problem_id,
                p.pdf_path,
                p.original_filename,
                p.updated_at.isoformat() if p.updated_at else "",
            ])
        zip_file.writestr("problem_submissions.csv", ps_out.getvalue().encode("utf-8-sig"))

        # 4. Problem States
        st_out = io.StringIO()
        st_writer = csv.writer(st_out)
        st_writer.writerow(["id", "problem_id", "title", "is_visible", "is_public_view", "deadline", "is_deleted", "created_at"])
        states = session.exec(select(ProblemState)).all()
        for s in states:
            st_writer.writerow([
                s.id,
                s.problem_id,
                s.title,
                s.is_visible,
                s.is_public_view,
                s.deadline.isoformat() if s.deadline else "",
                s.is_deleted,
                s.created_at.isoformat() if s.created_at else "",
            ])
        zip_file.writestr("problem_states.csv", st_out.getvalue().encode("utf-8-sig"))

        # 5. System Settings
        ss_out = io.StringIO()
        ss_writer = csv.writer(ss_out)
        ss_writer.writerow(["key", "value"])
        settings = session.exec(select(SystemSetting)).all()
        for s in settings:
            ss_writer.writerow([s.key, s.value])
        zip_file.writestr("system_settings.csv", ss_out.getvalue().encode("utf-8-sig"))

        # 6. TeamWork Configs
        twc_out = io.StringIO()
        twc_writer = csv.writer(twc_out)
        twc_writer.writerow(["id", "problem_id", "team_count", "team_size", "subproblem_count", "created_at", "updated_at"])
        team_configs = session.exec(select(TeamWorkConfig)).all()
        for c in team_configs:
            twc_writer.writerow([
                c.id,
                c.problem_id,
                c.team_count,
                c.team_size,
                c.subproblem_count,
                c.created_at.isoformat() if c.created_at else "",
                c.updated_at.isoformat() if c.updated_at else "",
            ])
        zip_file.writestr("teamwork_configs.csv", twc_out.getvalue().encode("utf-8-sig"))

        # 7. Teams
        t_out = io.StringIO()
        t_writer = csv.writer(t_out)
        t_writer.writerow(["id", "problem_id", "team_no", "name", "max_members", "created_at"])
        teams = session.exec(select(Team)).all()
        team_no_by_id = {t.id: t.team_no for t in teams}
        for t in teams:
            t_writer.writerow([
                t.id,
                t.problem_id,
                t.team_no,
                t.name,
                t.max_members,
                t.created_at.isoformat() if t.created_at else "",
            ])
        zip_file.writestr("teams.csv", t_out.getvalue().encode("utf-8-sig"))

        # 8. Team Members
        tm_out = io.StringIO()
        tm_writer = csv.writer(tm_out)
        tm_writer.writerow(["id", "problem_id", "team_id", "team_no", "student_id", "joined_at"])
        team_members = session.exec(select(TeamMember)).all()
        for m in team_members:
            tm_writer.writerow([
                m.id,
                m.problem_id,
                m.team_id,
                team_no_by_id.get(m.team_id),
                m.student_id,
                m.joined_at.isoformat() if m.joined_at else "",
            ])
        zip_file.writestr("team_members.csv", tm_out.getvalue().encode("utf-8-sig"))

        # 9. Team Claims
        tc_out = io.StringIO()
        tc_writer = csv.writer(tc_out)
        tc_writer.writerow(["id", "problem_id", "team_id", "team_no", "student_id", "subproblem_no", "claimed_at"])
        team_claims = session.exec(select(TeamSubproblemClaim)).all()
        for c in team_claims:
            tc_writer.writerow([
                c.id,
                c.problem_id,
                c.team_id,
                team_no_by_id.get(c.team_id),
                c.student_id,
                c.subproblem_no,
                c.claimed_at.isoformat() if c.claimed_at else "",
            ])
        zip_file.writestr("team_claims.csv", tc_out.getvalue().encode("utf-8-sig"))

        # 10. Team Attempts
        ta_out = io.StringIO()
        ta_writer = csv.writer(ta_out)
        ta_writer.writerow(["id", "problem_id", "team_id", "team_no", "student_id", "input_id", "attempts", "correct", "last_answer", "updated_at"])
        team_attempts = session.exec(select(TeamAttempt)).all()
        for a in team_attempts:
            ta_writer.writerow([
                a.id,
                a.problem_id,
                a.team_id,
            team_no_by_id.get(a.team_id),
                a.student_id,
                a.input_id,
                a.attempts,
                a.correct,
                a.last_answer,
                a.updated_at.isoformat() if a.updated_at else "",
            ])
        zip_file.writestr("team_attempts.csv", ta_out.getvalue().encode("utf-8-sig"))

        # 11. Team Submissions
        ts_out = io.StringIO()
        ts_writer = csv.writer(ts_out)
        ts_writer.writerow(["id", "problem_id", "team_id", "team_no", "student_id", "pdf_path", "original_filename", "updated_at"])
        team_submissions = session.exec(select(TeamSubmission)).all()
        for s in team_submissions:
            ts_writer.writerow([
                s.id,
                s.problem_id,
                s.team_id,
            team_no_by_id.get(s.team_id),
                s.student_id,
                s.pdf_path,
                s.original_filename,
                s.updated_at.isoformat() if s.updated_at else "",
            ])
        zip_file.writestr("team_submissions.csv", ts_out.getvalue().encode("utf-8-sig"))
        
        # 10. PDF Files
        submissions_dir = PUBLIC_DIR / "submissions"
        if submissions_dir.exists():
            for root, _, files in os.walk(submissions_dir):
                for file in files:
                    full_path = os.path.join(root, file)
                    rel_path = os.path.relpath(full_path, PUBLIC_DIR.parent)
                    zip_file.write(full_path, arcname=rel_path)
    
    return zip_buffer.getvalue()

def restore_db_from_bytes(content: bytes, session: Session):
    zip_buffer = io.BytesIO(content)
    try:
        with zipfile.ZipFile(zip_buffer, "r") as zf:
            names = zf.namelist()
            if "students.csv" not in names or "attempts.csv" not in names:
                raise ValueError("Missing CSV files in zip")

            def parse_bool(v):
                return str(v).lower() == 'true'

            def parse_dt(v):
                if not v:
                    return None
                return datetime.fromisoformat(v)
            
            # WIPE DATA
            session.exec(delete(TeamSubmission))
            session.exec(delete(TeamAttempt))
            session.exec(delete(TeamSubproblemClaim))
            session.exec(delete(TeamMember))
            session.exec(delete(Team))
            session.exec(delete(TeamWorkConfig))
            session.exec(delete(ProblemSubmission))
            session.exec(delete(Attempt))
            session.exec(delete(Student))

            if "problem_states.csv" in names:
                session.exec(delete(ProblemState))
            if "system_settings.csv" in names:
                session.exec(delete(SystemSetting))
            
            session.commit()
            
            # Restore Students
            with zf.open("students.csv") as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    obj = Student(
                        student_id=row['student_id'],
                        name=row['name'],
                        enabled=parse_bool(row['enabled']),
                        is_test=parse_bool(row['is_test']),
                        is_deleted=parse_bool(row['is_deleted']),
                        password_hash=row['password_hash'] if row['password_hash'] else None
                    )
                    session.add(obj)
            
            # Restore Attempts
            with zf.open("attempts.csv") as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                for row in reader:
                    obj = Attempt(
                        student_id=row['student_id'],
                        problem_id=row['problem_id'],
                        input_id=row['input_id'],
                        attempts=int(row['attempts']),
                        correct=parse_bool(row['correct']),
                        last_answer=row['last_answer'],
                        updated_at=parse_dt(row['updated_at']) or datetime.now()
                    )
                    session.add(obj)

            # Optional: Problem Submissions
            if "problem_submissions.csv" in names:
                with zf.open("problem_submissions.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = ProblemSubmission(
                            student_id=row['student_id'],
                            problem_id=row['problem_id'],
                            pdf_path=row['pdf_path'],
                            original_filename=row.get('original_filename') or None,
                            updated_at=parse_dt(row.get('updated_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Optional: Problem States
            if "problem_states.csv" in names:
                with zf.open("problem_states.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = ProblemState(
                            problem_id=row['problem_id'],
                            title=row.get('title') or None,
                            is_visible=parse_bool(row.get('is_visible')),
                            is_public_view=parse_bool(row.get('is_public_view')),
                            deadline=parse_dt(row.get('deadline')),
                            is_deleted=parse_bool(row.get('is_deleted')),
                            created_at=parse_dt(row.get('created_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Optional: System Settings
            if "system_settings.csv" in names:
                with zf.open("system_settings.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = SystemSetting(
                            key=row['key'],
                            value=row.get('value') or "",
                        )
                        session.add(obj)

            # Optional: TeamWork Configs
            if "teamwork_configs.csv" in names:
                with zf.open("teamwork_configs.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = TeamWorkConfig(
                            problem_id=row['problem_id'],
                            team_count=int(row['team_count']),
                            team_size=int(row['team_size']),
                            subproblem_count=int(row['subproblem_count']),
                            created_at=parse_dt(row.get('created_at')) or datetime.now(),
                            updated_at=parse_dt(row.get('updated_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Optional: Teams
            if "teams.csv" in names:
                with zf.open("teams.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = Team(
                            problem_id=row['problem_id'],
                            team_no=int(row['team_no']),
                            name=row.get('name') or None,
                            max_members=int(row.get('max_members') or 0),
                            created_at=parse_dt(row.get('created_at')) or datetime.now(),
                        )
                        session.add(obj)

            session.commit()

            # Build team id map from current DB rows to restore team members/attempts by logical key.
            team_rows = session.exec(select(Team)).all()
            team_id_map = {(t.problem_id, t.team_no): t.id for t in team_rows}

            def resolve_team_id(row):
                team_no_raw = row.get('team_no')
                if team_no_raw not in (None, ""):
                    try:
                        team_no = int(team_no_raw)
                        mapped = team_id_map.get((row['problem_id'], team_no))
                        if mapped is not None:
                            return int(mapped)
                    except (TypeError, ValueError):
                        pass

                team_id_raw = row.get('team_id')
                if team_id_raw in (None, ""):
                    raise ValueError(f"Missing team identity for problem {row.get('problem_id')}")
                return int(team_id_raw)

            # Optional: Team Members
            if "team_members.csv" in names:
                with zf.open("team_members.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        mapped_team_id = resolve_team_id(row)

                        obj = TeamMember(
                            problem_id=row['problem_id'],
                            team_id=int(mapped_team_id),
                            student_id=row['student_id'],
                            joined_at=parse_dt(row.get('joined_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Optional: Team Claims
            if "team_claims.csv" in names:
                with zf.open("team_claims.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = TeamSubproblemClaim(
                            problem_id=row['problem_id'],
                            team_id=resolve_team_id(row),
                            student_id=row['student_id'],
                            subproblem_no=int(row['subproblem_no']),
                            claimed_at=parse_dt(row.get('claimed_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Optional: Team Attempts
            if "team_attempts.csv" in names:
                with zf.open("team_attempts.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = TeamAttempt(
                            problem_id=row['problem_id'],
                            team_id=resolve_team_id(row),
                            student_id=row['student_id'],
                            input_id=row['input_id'],
                            attempts=int(row['attempts']),
                            correct=parse_bool(row['correct']),
                            last_answer=row.get('last_answer'),
                            updated_at=parse_dt(row.get('updated_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Optional: Team Submissions
            if "team_submissions.csv" in names:
                with zf.open("team_submissions.csv") as f:
                    reader = csv.DictReader(io.TextIOWrapper(f, encoding="utf-8-sig"))
                    for row in reader:
                        obj = TeamSubmission(
                            problem_id=row['problem_id'],
                            team_id=resolve_team_id(row),
                            student_id=row['student_id'],
                            pdf_path=row['pdf_path'],
                            original_filename=row.get('original_filename') or None,
                            updated_at=parse_dt(row.get('updated_at')) or datetime.now(),
                        )
                        session.add(obj)

            # Restore PDFs
            for name in names:
                if name.replace('\\', '/').startswith("public/submissions/"):
                    # The name is like public/submissions/pdf-test-problem/stu_pdf_test/file.pdf
                    file_data = zf.read(name)
                    # Create path inside backend folder
                    target_path = PUBLIC_DIR.parent / name
                    target_path.parent.mkdir(parents=True, exist_ok=True)
                    target_path.write_bytes(file_data)

            session.commit()
    except Exception as e:
        session.rollback()
        raise e

# --- Endpoints ---

@router.get("/db_dump")
def export_database(session: Session = Depends(get_session), token: str = Query(...)):
    verify_admin_token(token)
    zip_bytes = create_db_dump_bytes(session)
    response = Response(content=zip_bytes, media_type="application/zip")
    response.headers["Content-Disposition"] = f"attachment; filename=db_dump_{datetime.now().strftime('%Y%m%d')}.zip"
    return response

@router.post("/db_restore")
def import_database(
    file: UploadFile = File(...), 
    session: Session = Depends(get_session), 
    token: str = Query(...)
):
    verify_admin_token(token)
    if not file.filename.endswith('.zip'):
        raise HTTPException(status_code=400, detail="Only .zip files allowed")
    
    content = file.file.read()
    try:
        restore_db_from_bytes(content, session)
        return {"status": "success", "message": "Database restored"}
    except Exception as e:
        logging.error(f"Restore failed: {e}")
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")

# --- Managed Archive Endpoints ---

@router.post("/archives")
def create_archive(
    data: ArchiveCreateRequest,
    session: Session = Depends(get_session), 
    token: str = Query(...)
):
    """
    1. Dump current DB to ZIP.
    2. Save to archives folder with name.
    3. Wipe students (except test), attempts, and teamwork runtime records.
    """
    verify_admin_token(token)
    
    # Ensure Archives Directory
    if not os.path.exists(ARCHIVES_DIR):
        os.makedirs(ARCHIVES_DIR)
        
    # 1. Create Dump
    zip_bytes = create_db_dump_bytes(session)
    
    # 2. Save
    safe_name = "".join(c for c in data.name if c.isalnum() or c in (' ', '_', '-')).strip()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_{safe_name}.zip"
    file_path = os.path.join(ARCHIVES_DIR, filename)
    
    with open(file_path, "wb") as f:
        f.write(zip_bytes)
        
    # 3. Wipe Active Data (Semester Reset Logic)
    try:
        session.exec(delete(TeamSubmission))
        session.exec(delete(TeamAttempt))
        session.exec(delete(TeamSubproblemClaim))
        session.exec(delete(TeamMember))
        session.exec(delete(Team))
        session.exec(delete(ProblemSubmission))
        session.exec(delete(Attempt))
        
        # Delete non-test students
        statement = delete(Student).where(Student.is_test == False)
        session.exec(statement)
        
        # Optional: Reset problem states or keep them? 
        # Usually for a new semester, problems might stay same, just students change.
        # So we keep ProblemStates (assignments).
        
        session.commit()
    except Exception as e:
        # If wipe fails, we still have the backup, but we should warn
        logging.error(f"Archive created but wipe failed: {e}")
        return JSONResponse(status_code=500, content={"status": "partial_error", "message": f"Archive saved as {filename}, but DB wipe failed: {str(e)}"})
    
    return {"status": "success", "message": f"Archived to {filename} and database cleared."}

@router.get("/archives")
def list_archives(token: str = Query(...)):
    verify_admin_token(token)
    if not os.path.exists(ARCHIVES_DIR):
        return []
    
    files = []
    for f in os.listdir(ARCHIVES_DIR):
        if f.endswith(".zip"):
            path = os.path.join(ARCHIVES_DIR, f)
            stat = os.stat(path)
            files.append({
                "filename": f,
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
    
    # Sort by new
    files.sort(key=lambda x: x["filename"], reverse=True)
    return files

@router.post("/archives/restore/{filename}")
def restore_archive(
    filename: str,
    session: Session = Depends(get_session),
    token: str = Query(...)
):
    verify_admin_token(token)
    
    file_path = os.path.join(ARCHIVES_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archive not found")
        
    with open(file_path, "rb") as f:
        content = f.read()
        
    try:
        restore_db_from_bytes(content, session)
        return {"status": "success", "message": f"Restored from {filename}"}
    except Exception as e:
         raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")

@router.delete("/archives/{filename}")
def delete_archive(
    filename: str,
    token: str = Query(...)
):
    verify_admin_token(token)
    
    file_path = os.path.join(ARCHIVES_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archive not found")
        
    try:
        os.remove(file_path)
        return {"status": "success", "message": f"Deleted {filename}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")

# --- End New Endpoints ---

@router.get("/scores")
def export_scores(session: Session = Depends(get_session), token: str = Query(...)):
    verify_admin_token(token)
    students = session.exec(select(Student).where(Student.is_test == False, Student.is_deleted == False)).all()

    if not os.path.exists(PROBLEMS_DIR):
        problems = []
    else:
        published_ids = session.exec(select(ProblemState.problem_id).where(ProblemState.is_visible == True, ProblemState.is_deleted == False)).all()
        published_set = set(published_ids)
        problems = [f for f in os.listdir(PROBLEMS_DIR) if os.path.isdir(os.path.join(PROBLEMS_DIR, f)) and f in published_set]
        problems.sort()

    teamwork_configs = session.exec(select(TeamWorkConfig)).all()
    teamwork_config_map = {c.problem_id: c for c in teamwork_configs}

    problem_meta = {}
    for pid in problems:
        script = load_problem_script(pid)
        p_title = pid
        inputs_meta = {}
        if script and hasattr(script, "meta"):
            p_title = script.meta.get("title", pid)
            meta = ensure_meta_inputs(script.meta)
            inputs_meta = meta.get("inputs", {})

        # 这里必须以题目真实 input 列表为准，不能只看 meta.inputs。
        # 否则 meta 未完整声明时，总分母会错误地变成 0，导致各种得分率显示 0。
        input_ids = get_problem_input_ids(pid, script)

        input_items = []
        input_scores = {}
        total_possible = 0
        for iid in sorted(set(input_ids)):
            conf = inputs_meta.get(iid, {}) if isinstance(inputs_meta, dict) else {}
            score = conf.get("score", DEFAULT_INPUT_SCORE)
            input_items.append((iid, score))
            input_scores[iid] = score
            total_possible += score

        teamwork_cfg = teamwork_config_map.get(pid)
        input_sub_map = {}
        sub_total_scores = {}
        if teamwork_cfg and script:
            rng = get_stable_rng(f"public_{pid}")
            try:
                params = script.generate(rng) if hasattr(script, "generate") else {}
            except Exception:
                params = {}
            bundle = get_problem_subproblem_bundle(pid, params)
            input_sub_map = bundle.get("input_subproblem_map", {})
            for iid, score in input_items:
                sub_no = input_sub_map.get(iid)
                if sub_no is None:
                    continue
                sub_total_scores[sub_no] = sub_total_scores.get(sub_no, 0) + score

        problem_meta[pid] = {
            "title": p_title,
            "is_teamwork": bool(teamwork_cfg),
            "input_items": input_items,
            "input_scores": input_scores,
            "total_possible": total_possible,
            "input_sub_map": input_sub_map,
            "sub_total_scores": sub_total_scores,
        }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "scores"

    # 生成双层表头：
    # 第 1 行是“题目分组(标题+ID)”，第 2 行是“具体字段(总分/得分率/input_id)”。
    # 同时在每道题后插入空列作为视觉分隔，便于在 Excel 中快速区分。
    header_top = ["基础信息", "基础信息", "基础信息"]
    header_sub = ["学号", "姓名", "总分"]

    for idx, pid in enumerate(problems):
        p_meta = problem_meta.get(pid, {})
        p_title = p_meta.get("title", pid)

        sub_headers = ["总分"]
        if p_meta.get("is_teamwork"):
            sub_headers.extend(["个人得分率%", "队伍总分", "队伍得分率%"])
        for iid, _ in p_meta.get("input_items", []):
            sub_headers.append(iid)

        group_label = f"{p_title} [{pid}]"
        header_top.extend([group_label] * len(sub_headers))
        header_sub.extend(sub_headers)

        # 题目间加一列空白分隔（最后一题不加）
        if idx < len(problems) - 1:
            header_top.append("")
            header_sub.append("")

    attempts = session.exec(select(Attempt)).all()
    att_map = {}
    for a in attempts:
        att_map[(a.student_id, a.problem_id, a.input_id)] = a

    team_attempts = session.exec(select(TeamAttempt)).all()
    team_att_map = {}
    for a in team_attempts:
        team_att_map[(a.student_id, a.problem_id, a.input_id)] = a

    team_members = session.exec(select(TeamMember)).all()
    member_team_map = {(m.problem_id, m.student_id): m.team_id for m in team_members}

    team_claims = session.exec(select(TeamSubproblemClaim)).all()
    claim_map = {(c.problem_id, c.team_id, c.student_id): c.subproblem_no for c in team_claims}

    team_seen_inputs = set()
    team_score_map = {}
    for a in team_attempts:
        if not a.correct:
            continue

        p_meta = problem_meta.get(a.problem_id)
        if not p_meta or not p_meta.get("is_teamwork"):
            continue

        key = (a.problem_id, a.team_id, a.input_id)
        if key in team_seen_inputs:
            continue
        team_seen_inputs.add(key)

        score = p_meta.get("input_scores", {}).get(a.input_id, DEFAULT_INPUT_SCORE)
        team_score_map[(a.problem_id, a.team_id)] = team_score_map.get((a.problem_id, a.team_id), 0) + score

    # 团队题在提交时使用 student_id 作为随机种子的一部分，
    # 导出 CSV 时也必须使用相同种子重建 input->subproblem 映射，否则会把正确题目误判成 0 分。
    student_team_input_sub_map_cache = {}

    def get_student_team_input_sub_map(student_id: str, problem_id: str):
        cache_key = (student_id, problem_id)
        if cache_key in student_team_input_sub_map_cache:
            return student_team_input_sub_map_cache[cache_key]

        p_meta = problem_meta.get(problem_id, {})
        default_map = p_meta.get("input_sub_map", {})

        script = load_problem_script(problem_id)
        if not script:
            student_team_input_sub_map_cache[cache_key] = default_map
            return default_map

        rng = get_stable_rng(f"{student_id}_{problem_id}")
        try:
            params = script.generate(rng) if hasattr(script, "generate") else {}
        except Exception:
            params = {}

        bundle = get_problem_subproblem_bundle(problem_id, params)
        input_sub_map = bundle.get("input_subproblem_map", {}) or default_map
        student_team_input_sub_map_cache[cache_key] = input_sub_map
        return input_sub_map

    student_rows = []
    for s in students:
        row = [s.student_id, s.name, 0]
        current_total = 0

        for idx, pid in enumerate(problems):
            p_meta = problem_meta.get(pid, {})
            is_teamwork = p_meta.get("is_teamwork", False)

            p_total_idx = len(row)
            row.append(0)

            personal_rate_idx = None
            team_total_idx = None
            team_rate_idx = None
            if is_teamwork:
                personal_rate_idx = len(row)
                row.append(0)
                team_total_idx = len(row)
                row.append(0)
                team_rate_idx = len(row)
                row.append(0)

            p_total = 0
            team_id = member_team_map.get((pid, s.student_id)) if is_teamwork else None
            claim_sub = claim_map.get((pid, team_id, s.student_id)) if is_teamwork and team_id else None

            student_input_sub_map = {}
            if is_teamwork:
                student_input_sub_map = get_student_team_input_sub_map(s.student_id, pid)

            for iid, max_val in p_meta.get("input_items", []):
                score = 0
                if is_teamwork:
                    mapped_sub = student_input_sub_map.get(iid)
                    if claim_sub is not None and mapped_sub == claim_sub:
                        att = team_att_map.get((s.student_id, pid, iid))
                        if att and att.correct:
                            score = max_val
                else:
                    att = att_map.get((s.student_id, pid, iid))
                    if att and att.correct:
                        score = max_val

                row.append(score)
                p_total += score

            row[p_total_idx] = p_total
            current_total += p_total

            if is_teamwork:
                if claim_sub is not None:
                    personal_possible = sum(
                        score
                        for iid, score in p_meta.get("input_items", [])
                        if student_input_sub_map.get(iid) == claim_sub
                    )
                else:
                    personal_possible = 0
                personal_rate = round((p_total / personal_possible) * 100, 1) if personal_possible > 0 else 0
                team_total = team_score_map.get((pid, team_id), 0) if team_id else 0
                prob_total_possible = p_meta.get("total_possible", 0)
                team_rate = round((team_total / prob_total_possible) * 100, 1) if prob_total_possible > 0 else 0

                row[personal_rate_idx] = personal_rate
                row[team_total_idx] = team_total
                row[team_rate_idx] = team_rate

            # 与表头一致，题目之间插入空白分隔列
            if idx < len(problems) - 1:
                row.append("")

        row[2] = current_total
        student_rows.append(row)

    # 转置为“学生为列”的布局：
    # 第1行放学号（列标题），第2行放姓名，后续每行是一个指标。
    # 题目分组单独放在第1列，字段名放在第2列，并对连续相同分组合并。
    sheet.delete_rows(1, sheet.max_row)

    student_ids = [r[0] for r in student_rows]
    student_names = [r[1] for r in student_rows]

    sheet.append(["题目分组", "字段"] + student_ids)
    sheet.append(["", "姓名"] + student_names)

    # 从“总分”列开始转置（跳过原来的学号/姓名列）
    for col_idx in range(2, len(header_sub)):
        group_label = header_top[col_idx]
        field_label = header_sub[col_idx]

        # 保留题目间隔行，提升可读性
        if not field_label and not group_label:
            sheet.append(["", ""] + [""] * len(student_rows))
            continue

        values = [r[col_idx] if col_idx < len(r) else "" for r in student_rows]
        sheet.append([group_label, field_label] + values)

    # 合并第1列连续重复分组（从第3行开始，避免合并表头）
    merge_start = None
    merge_label = None
    for row_idx in range(3, sheet.max_row + 2):
        current_label = sheet.cell(row=row_idx, column=1).value if row_idx <= sheet.max_row else None
        if current_label and current_label == merge_label:
            continue

        if merge_label and merge_start is not None and row_idx - 1 > merge_start:
            sheet.merge_cells(start_row=merge_start, start_column=1, end_row=row_idx - 1, end_column=1)

        if current_label:
            merge_start = row_idx
            merge_label = current_label
        else:
            merge_start = None
            merge_label = None

    # 表头样式
    for row_idx in (1, 2):
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=c)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    # 冻结前两行和前两列，便于查看
    sheet.freeze_panes = "C3"

    # 简单自动列宽
    for col_idx, col_cells in enumerate(sheet.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    output = io.BytesIO()
    workbook.save(output)
    content_bytes = output.getvalue()
    
    response = Response(
        content=content_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f"attachment; filename=scores_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return response

@router.get("/work")
def export_work(session: Session = Depends(get_session), token: str = Query(...)):
    """
    导出所有学生的作业（打包为 ZIP 文件）。
    ZIP 文件中，每个学生一个文件夹，文件夹内包含：
    1. 该学生每道题答题情况的 Markdown 文件（包含作答记录和得分）
    2. 学生额外上传的附件 PDF（如有）
    注：为了降低服务器性能压力，总归档 PDF 的生成已移除，改为将源文件直接导出。
    """
    verify_admin_token(token)
    
    # 1. 获取需要统计的学生列表（排除测试账号和已删除账号）
    students = session.exec(select(Student).where(Student.is_test == False, Student.is_deleted == False)).all()
    
    if not os.path.exists(PROBLEMS_DIR):
        return Response("No problems found", status_code=404)
    
    # 2. 获取所有已发布且未被删除的题目
    published_ids = session.exec(select(ProblemState.problem_id).where(ProblemState.is_visible == True, ProblemState.is_deleted == False)).all()
    published_set = set(published_ids)
        
    problems = [f for f in os.listdir(PROBLEMS_DIR) if os.path.isdir(os.path.join(PROBLEMS_DIR, f)) and f in published_set]
    problems.sort()
    
    # 将打包生成的 ZIP 文件流缓存在系统内存中
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # =====================================================================
        # 3. 提前批量加载数据库记录，建立内存索引字典，避免在双层大循环中频繁查库(解决 N+1 查询问题)
        # =====================================================================
        
        # 建立题目对应的团队作业配置映射
        teamwork_configs = session.exec(select(TeamWorkConfig)).all()
        teamwork_config_map = {c.problem_id: c for c in teamwork_configs}

        # 缓存所有学生的【团队作业】答题记录
        all_team_attempts = session.exec(select(TeamAttempt)).all()
        team_attempt_map = {}
        for a in all_team_attempts:
            if a.student_id not in team_attempt_map:
                team_attempt_map[a.student_id] = {}
            if a.problem_id not in team_attempt_map[a.student_id]:
                team_attempt_map[a.student_id][a.problem_id] = {}
            team_attempt_map[a.student_id][a.problem_id][a.input_id] = a

        # 缓存学生的团队归属情况
        all_team_members = session.exec(select(TeamMember)).all()
        member_team_map = {(m.problem_id, m.student_id): m.team_id for m in all_team_members}

        # 缓存团队作业中，学生已认领的具体子题目(板块)编号
        all_claims = session.exec(select(TeamSubproblemClaim)).all()
        claim_map = {(c.problem_id, c.team_id, c.student_id): c.subproblem_no for c in all_claims}

        # 缓存所有普通作业答题记录
        all_attempts = session.exec(select(Attempt)).all()
        mega_map = {}
        for a in all_attempts:
            if a.student_id not in mega_map: mega_map[a.student_id] = {}
            if a.problem_id not in mega_map[a.student_id]: mega_map[a.student_id][a.problem_id] = {}
            mega_map[a.student_id][a.problem_id][a.input_id] = a

        # 缓存学生额外上传的文件附录（普通作业及团队作业相关附件）
        all_prob_subs = session.exec(select(ProblemSubmission)).all()
        all_team_subs = session.exec(select(TeamSubmission)).all()
        sub_folder_map = {}
        for sub in all_prob_subs:
            sub_folder_map[(sub.student_id, sub.problem_id)] = sub
        for sub in all_team_subs:
            sub_folder_map[(sub.student_id, sub.problem_id)] = sub

        # =====================================================================
        # 4. 遍历所有学生及其题目，生成静态导出文件并在内存中写入 ZIP 包中
        # =====================================================================
        for s in students:
            # 建立学生专属路径格式，如: "2024001_张三"
            s_folder = f"{s.student_id}_{s.name}"

            # 遍历每道已发布的题目
            for pid in problems:
                script = load_problem_script(pid)
                if not script: continue
                
                # 初始化基于“学号_题目ID”的固定随机数种子（确保与学生在线做题时的参数生成的完全一致）
                rng = get_stable_rng(f"{s.student_id}_{pid}")
                try:
                    params = script.generate(rng)
                except:
                    params = {}
                
                md_path = os.path.join(PROBLEMS_DIR, pid, "problem.md")
                if not os.path.exists(md_path): continue
                
                with open(md_path, "r", encoding="utf-8") as f:
                    content = f.read()

                # 读取并处理满分配分配置及元数据
                if hasattr(script, "meta"):
                    meta = ensure_meta_inputs(script.meta)
                    full_meta_inputs = meta.get("inputs", {})
                else:
                    full_meta_inputs = {}

                # =============================================================
                # A: 评分与答题数据过滤逻辑（专门区分团队作业与普通作业的不同）
                # =============================================================
                teamwork_cfg = teamwork_config_map.get(pid)
                if teamwork_cfg:
                    # 如果是团队作业：获取该学生在此团队作业下的原始作答记录
                    s_attempts_all = team_attempt_map.get(s.student_id, {}).get(pid, {})

                    # 判断所属队伍以及认领的具体子板块
                    team_id = member_team_map.get((pid, s.student_id))
                    claim_sub = claim_map.get((pid, team_id, s.student_id)) if team_id else None

                    # 获得整道大题和具体每个填空框(input_id)对应子板块序号的映射关系
                    bundle = get_problem_subproblem_bundle(pid, params)
                    input_sub_map = bundle.get("input_subproblem_map", {})

                    # 把学生真正认领板块里面的包含的所有空都筛选出来
                    allowed_input_ids = set()
                    if claim_sub is not None:
                        for iid, sub_no in input_sub_map.items():
                            if sub_no == claim_sub:
                                allowed_input_ids.add(iid)

                    # 过滤只剩下“其负责部分”的真实答题记录
                    s_attempts = {iid: rec for iid, rec in s_attempts_all.items() if iid in allowed_input_ids}

                    # 我们保持把队伍作业的所有空都带上(以便在最终报告中渲染出整份试卷)，
                    # 但是对于不属于该学生认领的空，我们要将其该空的配分强制设为 0，以此规避错误统计其他人的分
                    meta_inputs = {}
                    for iid, conf in full_meta_inputs.items():
                        conf_dict = dict(conf) if isinstance(conf, dict) else {}
                        if iid not in allowed_input_ids:
                            conf_dict["score"] = 0
                        meta_inputs[iid] = conf_dict
                else:
                    # 如果是普通作业：获取全部填空配置及学生个人的历史答复（无阉割）
                    s_attempts = mega_map.get(s.student_id, {}).get(pid, {})
                    meta_inputs = full_meta_inputs
                
                # =============================================================
                # B: 利用 Jinja2 渲染动态参数并注入包含样式（学生分数和答案）的文本
                # =============================================================
                # 初始化专门处理 {{ input("ans_1") }} 的助手类 ExportInputHelper
                helper = ExportInputHelper(s_attempts, meta_inputs)
                render_context = {**params, "input": helper}

                # 使用 Jinja2 执行替换
                template = Template(content)
                filled_md = template.render(**render_context)
                
                # 汇总计算该学生在这道题目（部分）上获得的总得分
                summary = f"\n\n---\n**总得分: {helper.total_obtained} / {helper.total_possible}**"
                filled_md += summary
                final_md = f"# {pid}\n\n{filled_md}"
                
                # 将学生此题的答题反馈放入 ZIP 的以该学生名字命名的文件夹中
                zip_file.writestr(f"{s_folder}/{pid}.md", final_md)

                # =============================================================
                # C: 读取题目的附件(例如 PDF 版手写步骤)，统一归并到打入该 ZIP 中
                # =============================================================
                submission = sub_folder_map.get((s.student_id, pid))
                if submission and submission.pdf_path:
                    pdf_full_path = (PUBLIC_DIR / submission.pdf_path).resolve()
                    if pdf_full_path.exists() and pdf_full_path.is_file():
                        pdf_name = submission.original_filename or os.path.basename(submission.pdf_path)
                        zip_file.write(pdf_full_path, arcname=f"{s_folder}/{pid}_附件_{pdf_name}")

    zip_buffer.seek(0)
    response = Response(content=zip_buffer.getvalue(), media_type="application/zip")
    response.headers["Content-Disposition"] = f"attachment; filename=student_work_{datetime.now().strftime('%Y%m%d')}.zip"
    return response
